"""
Génère un Excel avec les vraies formules Excel :
  - =SMALL($F$4:$F$N, k)         → VaR
  - =AVERAGE($H$4:INDEX(...,k-1)) → CVaR
  - =FLOOR(J4*J5,1)              → k
  - =SMALL($F$4:$F$N, ROW()-3)   → colonne triée
  - =C{r-1}*(1+B{r})             → valeur portefeuille cumulée
  - =C{r}/$C$3-1                 → rendement cumulé
AI.PA (50%) + AIR.PA (50%) — q=5% — 100 000 € — fin 12/02/2026 — perf 04/03/2021
"""
import sys, os, math
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding="utf-8")

from app.services.data_fetcher import _fetch_prices_sync
import pandas as pd, numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paramètres ────────────────────────────────────────────────────────────────
TICKERS    = ["AI.PA", "AIR.PA"]
WEIGHTS    = {"AI.PA": 0.50, "AIR.PA": 0.50}
NAMES      = {"AI.PA": "Air Liquide", "AIR.PA": "Airbus"}
END_DATE   = "2026-02-12"
PERF_START = "2021-03-04"
VAR_START  = "2006-02-12"   # date_fin − 20 ans
QUANTILE   = 0.05
AMOUNT     = 100_000

# ── Fetch ─────────────────────────────────────────────────────────────────────
print("Téléchargement des données...")
prices, errors = _fetch_prices_sync(TICKERS, "2000-01-03", END_DATE)
if prices.empty: print("ERREUR"); sys.exit(1)
print(f"OK : {len(prices)} lignes.")

# ── Calculs de référence ─────────────────────────────────────────────────────
ret_all  = prices[TICKERS].pct_change().dropna()
ret_all  = ret_all[(ret_all.abs() <= 0.40).all(axis=1)]

ret_var  = ret_all[ret_all.index >= pd.Timestamp(VAR_START)]
ret_perf = ret_all[ret_all.index >= pd.Timestamp(PERF_START)]

w_arr = np.array([WEIGHTS[t] for t in TICKERS])
pr_var  = ret_var.dot(w_arr)
pr_perf = ret_perf.dot(w_arr)

sv = np.sort(pr_var.values)
n  = len(sv)
k  = max(1, int(math.floor(n * QUANTILE)))
var_pct = sv[k-1]
cvar_pct = sv[:k-1].mean() if k > 1 else var_pct

port_vol  = float(pr_perf.std() * np.sqrt(252)) * 100
port_mean = float(pr_perf.mean() * 252) * 100
tot_ret   = float((1 + pr_perf).cumprod().iloc[-1] - 1) * 100

print(f"Fenêtre VaR : {ret_var.index[0].date()} → {ret_var.index[-1].date()}  ({n} obs)")
print(f"k={k}  VaR={var_pct*100:.4f}%  CVaR={cvar_pct*100:.4f}%")

# ── Styles ────────────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="BBBBBB")
BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR  = Alignment(horizontal="center", vertical="center")
LFT  = Alignment(horizontal="left",   vertical="center")
RGT  = Alignment(horizontal="right",  vertical="center")

def fill(hex): return PatternFill("solid", fgColor=hex)
F_DBLUE  = fill("1F4E79"); F_RED    = fill("C00000"); F_LRED   = fill("FADBD8")
F_GREEN  = fill("375623"); F_LBLUE  = fill("D6E4F0"); F_LGREEN = fill("E8F5E9")
F_ORANGE = fill("FFF2CC"); F_LGRAY  = fill("F2F2F2")

def font(color="000000", bold=False, size=10, italic=False):
    return Font(color=color, bold=bold, size=size, italic=italic)
FW = font("FFFFFF", bold=True)
FR = font("C00000", bold=True)
FG = font("375623", bold=True)
FB = font("1F4E79", bold=True)

def hdr(ws, r, c, text):
    cell = ws.cell(row=r, column=c, value=text)
    cell.fill = F_DBLUE; cell.font = FW
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BRD; return cell

def vc(ws, r, c, v, fmt=None, fi=None, fo=None, al=CTR):
    cell = ws.cell(row=r, column=c, value=v)
    if fmt: cell.number_format = fmt
    if fi:  cell.fill = fi
    if fo:  cell.font = fo
    cell.alignment = al; cell.border = BRD; return cell

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# FEUILLE 1 — VaR 95% 1J
# Structure :
#   A=Date  B=Prix AI.PA  C=Prix AIR.PA  D=Rdt AI.PA  E=Rdt AIR.PA
#   F=Rdt Portefeuille (formule)  [G=vide]  H=Rdts triés (formule SMALL)
#   [vide]  I=Libellé  J=Valeur/Formule
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "VaR 95% 1J"

# largeurs colonnes
for col, w in zip("ABCDEFGHIJ", [13,14,14,13,13,16,4,18,36,22]):
    ws1.column_dimensions[col].width = w

DR = 4          # première ligne de données
n_r = len(ret_var)
LR = DR + n_r - 1  # dernière ligne de données

# Plages utiles
F_RNG = f"$F${DR}:$F${LR}"   # rendements portefeuille
H_RNG = f"$H${DR}:$H${LR}"   # rendements triés

# Lignes métriques dans col J
RN   = DR      # n
RQ   = DR+1    # q
RK   = DR+2    # k
RVAR = DR+4    # VaR %
RVAE = DR+5    # VaR €
RCVR = DR+7    # CVaR %
RCVE = DR+8    # CVaR €

# ── Titre ────────────────────────────────────────────────────────────────────
ws1.merge_cells("A1:J1")
ws1["A1"].value = (f"VaR Historique 95% 1J — AI.PA 50% + AIR.PA 50% — "
                   f"Fenêtre VaR : {ret_var.index[0].date()} → {ret_var.index[-1].date()}")
ws1["A1"].font = font("1F4E79", bold=True, size=12)
ws1["A1"].alignment = CTR; ws1.row_dimensions[1].height = 26

ws1.merge_cells("A2:H2")
ws1["A2"].value = ("Rendements journaliers filtrés (|r| ≤ 40%)  —  "
                   "Rdt = (Pᵢ − Pᵢ₋₁) / Pᵢ₋₁  —  "
                   "Rdt portefeuille = 0,5 × rAI + 0,5 × rAIR")
ws1["A2"].font = font("666666", italic=True, size=9)
ws1["A2"].alignment = LFT; ws1.row_dimensions[2].height = 16

# ── En-têtes colonnes ────────────────────────────────────────────────────────
ws1.row_dimensions[3].height = 36
for c, txt in [
    (1,"Date"), (2,"AI.PA\ncours"), (3,"AIR.PA\ncours"),
    (4,"Rdt AI.PA\n(valeur filtrée)"), (5,"Rdt AIR.PA\n(valeur filtrée)"),
    (6,"Rdt Portefeuille\n=0,5×D+0,5×E"),
    (8,"Rdts triés ↑\n=SMALL($F,i)"),
    (9,"Indicateur"), (10,"Valeur / Formule"),
]:
    hdr(ws1, 3, c, txt)

# ── Panneau métriques (col I/J) ───────────────────────────────────────────────
def metric(r, lbl, formula_or_val, fmt=None, is_key=False):
    fi = F_LRED if is_key else None
    fo = FR     if is_key else None
    vc(ws1, r, 9,  lbl,             fi=fi, al=LFT)
    vc(ws1, r, 10, formula_or_val,  fmt=fmt, fi=fi, fo=fo)

metric(RN,   "n = COUNTA(rdts portefeuille)",         f"=COUNTA({F_RNG})",         "0")
metric(RQ,   "q = quantile choisi",                    QUANTILE,                    "0%")
metric(RK,   "k = FLOOR(n × q, 1)",                   f"=FLOOR(J{RN}*J{RQ},1)",   "0", is_key=True)
# ligne vide RK+1
vc(ws1, RK+1, 9, ""); vc(ws1, RK+1, 10, "")

metric(RVAR, f"VaR 95% (%) = SMALL(F, k)",             f"=SMALL({F_RNG},J{RK})",    "0.0000%",   is_key=True)
metric(RVAE, f"VaR 95% (€) = |VaR%| × {AMOUNT:,}",   f"=ABS(J{RVAR})*{AMOUNT}",  '#,##0.00 "€"', is_key=True)
# ligne vide RVAE+1
vc(ws1, RVAE+1, 9, ""); vc(ws1, RVAE+1, 10, "")

# CVaR = AVERAGE des k-1 premiers rdts triés
cvar_formula = f"=AVERAGE($H${DR}:INDEX({H_RNG},J{RK}-1))"
metric(RCVR, "CVaR (%) = AVERAGE(H[1..k−1])",          cvar_formula,                "0.0000%")
metric(RCVE, f"CVaR (€) = |CVaR%| × {AMOUNT:,}",      f"=ABS(J{RCVR})*{AMOUNT}",  '#,##0.00 "€"')

# Infos perf
vc(ws1, RCVE+2, 9, "─── Fenêtre performance ───")
ws1.cell(RCVE+2, 9).font = font("888888", italic=True, size=9)
metric(RCVE+3, "Volatilité annualisée (σ × √252)", port_vol/100,  "0.00%")
metric(RCVE+4, "Rendement total (fenêtre perf)",   tot_ret/100,   "0.00%")

# Légende couleurs
vc(ws1, RCVE+6, 9, "Rouge foncé = VaR (k-ième pire)")
ws1.cell(RCVE+6, 9).fill = F_RED; ws1.cell(RCVE+6, 9).font = FW
vc(ws1, RCVE+7, 9, "Rose = CVaR (k−1 pires rdts)")
ws1.cell(RCVE+7, 9).fill = F_LRED

# ── Données ───────────────────────────────────────────────────────────────────
prices_w = prices[TICKERS]
prices_w = prices_w[(prices_w.index >= pd.Timestamp(VAR_START)) &
                    (prices_w.index <= pd.Timestamp(END_DATE))]
ret_w    = prices_w.pct_change().dropna()
ret_w    = ret_w[(ret_w.abs() <= 0.40).all(axis=1)]

for i, (dt, row) in enumerate(ret_w.iterrows()):
    r = DR + i

    # Date
    c = ws1.cell(r, 1, dt.date()); c.number_format = "DD/MM/YYYY"
    c.border = BRD; c.alignment = CTR

    # Prix
    for j, t in enumerate(TICKERS, 2):
        p = prices_w.loc[dt, t] if dt in prices_w.index else None
        vc(ws1, r, j, p, fmt='#,##0.00', al=RGT)

    # Rendements individuels (valeurs filtrées)
    for j, t in enumerate(TICKERS, 4):
        vc(ws1, r, j, float(row[t]), fmt="0.0000%", al=RGT)

    # Rendement portefeuille — FORMULE
    c = ws1.cell(r, 6, f"={WEIGHTS['AI.PA']}*D{r}+{WEIGHTS['AIR.PA']}*E{r}")
    c.number_format = "0.0000%"; c.border = BRD; c.alignment = RGT

    # Colonne triée — FORMULE =SMALL($F$DR:$F$LR, i)
    rank = i + 1
    c = ws1.cell(r, 8, f"=SMALL({F_RNG},{rank})")
    c.number_format = "0.0000%"; c.border = BRD; c.alignment = RGT
    if rank == k:                            # k-ième = VaR
        c.fill = F_RED; c.font = font("FFFFFF", bold=True)
    elif rank < k:                           # k-1 premiers = CVaR
        c.fill = F_LRED

# Figer le volet
ws1.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════════
# FEUILLE 2 — Performance
# Formules : =C{r-1}*(1+B{r})  et  =C{r}/$C$3-1
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Performance")
for col, w in zip("ABCDE", [13, 15, 22, 14, 14]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells("A1:E1")
ws2["A1"].value = (f"Performance du portefeuille — "
                   f"{ret_perf.index[0].date()} → {ret_perf.index[-1].date()}")
ws2["A1"].font = font("1F4E79", bold=True, size=12); ws2["A1"].alignment = CTR
ws2.row_dimensions[1].height = 26

for j, h in enumerate(["Date","Rdt journalier",
                        f"Valeur ({AMOUNT:,.0f} €)","Rdt cumulé","Rdt cumulé %"], 1):
    hdr(ws2, 2, j, h)
ws2.row_dimensions[2].height = 30

# Ligne initiale (row 3)
ws2.cell(3,1).value = "Valeur initiale"; ws2.cell(3,1).border = BRD
vc(ws2, 3, 2,  0,      fmt="0.0000%")
vc(ws2, 3, 3,  AMOUNT, fmt='#,##0.00 "€"', fo=FB)
vc(ws2, 3, 4,  0,      fmt="0.00%")
vc(ws2, 3, 5,  0,      fmt='0.00"%"')

for i, (dt, dr) in enumerate(pr_perf.items()):
    r = 4 + i
    c1 = ws2.cell(r, 1, dt.date()); c1.number_format = "DD/MM/YYYY"
    c1.border = BRD; c1.alignment = CTR

    # Rendement journalier (valeur)
    vc(ws2, r, 2, float(dr), fmt="0.0000%",
       fo=font("375623" if dr >= 0 else "C00000"), al=RGT)

    # Valeur cumulée — FORMULE chaînée
    c3 = ws2.cell(r, 3, f"=C{r-1}*(1+B{r})")
    c3.number_format = '#,##0.00 "€"'; c3.border = BRD; c3.alignment = RGT

    # Rendement cumulé — FORMULE
    c4 = ws2.cell(r, 4, f"=C{r}/$C$3-1")
    c4.number_format = "0.00%"; c4.border = BRD; c4.alignment = RGT

    c5 = ws2.cell(r, 5, f"=(C{r}/$C$3-1)*100")
    c5.number_format = '0.00"%"'; c5.border = BRD; c5.alignment = RGT

ws2.freeze_panes = "A3"
LAST_PERF = 3 + len(pr_perf)

# ═══════════════════════════════════════════════════════════════════════════════
# FEUILLE 3 — Corrélation & Covariance
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Corrélation")
ws3.column_dimensions["A"].width = 26
for c in ["B","C"]: ws3.column_dimensions[c].width = 18

ws3.merge_cells("A1:C1")
ws3["A1"].value = "Matrice de corrélation & covariance — Fenêtre performance"
ws3["A1"].font = font("1F4E79", bold=True, size=12); ws3["A1"].alignment = CTR

corr = ret_perf[TICKERS].corr()
cov  = ret_perf[TICKERS].cov()

ws3.cell(3,1).value = "Corrélation (ρ)"; ws3.cell(3,1).font = font(bold=True)
for j, t in enumerate(TICKERS, 2): hdr(ws3, 3, j, NAMES[t])
for i, t1 in enumerate(TICKERS, 4):
    hdr(ws3, i, 1, NAMES[t1])
    for j, t2 in enumerate(TICKERS, 2):
        vc(ws3, i, j, round(float(corr.loc[t1,t2]),6), "0.0000")

ws3.cell(7,1).value = "Covariance quotidienne (σ₁σ₂ρ)"; ws3.cell(7,1).font = font(bold=True)
for j, t in enumerate(TICKERS, 2): hdr(ws3, 7, j, NAMES[t])
for i, t1 in enumerate(TICKERS, 8):
    hdr(ws3, i, 1, NAMES[t1])
    for j, t2 in enumerate(TICKERS, 2):
        vc(ws3, i, j, round(float(cov.loc[t1,t2]),8), "0.000000")

# Formule volatilité portefeuille (cours)
s1  = float(ret_perf["AI.PA"].std())
s2  = float(ret_perf["AIR.PA"].std())
rho = float(corr.loc["AI.PA","AIR.PA"])
w1, w2 = WEIGHTS["AI.PA"], WEIGHTS["AIR.PA"]
vol_ann = math.sqrt((w1**2*s1**2 + w2**2*s2**2 + 2*w1*w2*rho*s1*s2)*252)*100

ws3.merge_cells("A11:C11")
ws3["A11"].value = "Formule volatilité portefeuille (cours M2) :"
ws3["A11"].font  = font("1F4E79", bold=True)

ws3.merge_cells("A12:C12")
ws3["A12"].value = "σₚ = √( w₁²σ₁² + w₂²σ₂² + 2·w₁·w₂·ρ·σ₁·σ₂ ) × √252"
ws3["A12"].font  = font(bold=True); ws3["A12"].alignment = LFT

for i, (lbl, v, fmt) in enumerate([
    ("σ₁ quotidien AI.PA",   s1,   "0.000000%"),
    ("σ₂ quotidien AIR.PA",  s2,   "0.000000%"),
    ("ρ corrélation",         rho,  "0.0000"),
    ("w₁",                    w1,   "0%"),
    ("w₂",                    w2,   "0%"),
], 13):
    ws3.cell(i,1).value = lbl; ws3.cell(i,1).fill = F_LBLUE
    vc(ws3, i, 2, v, fmt=fmt)

ws3.cell(18,1).value = "σₚ annualisée"; ws3.cell(18,1).font = font(bold=True)
vc(ws3, 18, 2, vol_ann/100, "0.00%", fo=FB)

# ═══════════════════════════════════════════════════════════════════════════════
# FEUILLE 4 — Résumé (références vers les autres feuilles)
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Résumé")
ws4.column_dimensions["A"].width = 42
ws4.column_dimensions["B"].width = 28
ws4.column_dimensions["C"].width = 20

ws4.merge_cells("A1:C1")
ws4["A1"].value = "VaR Historique 1 Jour — Synthèse"
ws4["A1"].font  = font("1F4E79", bold=True, size=14); ws4["A1"].alignment = CTR
ws4.row_dimensions[1].height = 32

# Paramètres
ws4.merge_cells("A3:C3")
ws4["A3"].value = "PARAMÈTRES"; ws4["A3"].fill = F_DBLUE
ws4["A3"].font  = font("FFFFFF", bold=True, size=11); ws4["A3"].alignment = LFT

for i, (lbl, v1, v2) in enumerate([
    ("Action 1",                       "Air Liquide (AI.PA)",                                   "50%"),
    ("Action 2",                       "Airbus (AIR.PA)",                                        "50%"),
    ("Montant du portefeuille",        f"{AMOUNT:,.0f} €",                                       ""),
    ("Quantile q",                     f"{QUANTILE*100:.0f}%",                                   ""),
    ("Niveau de confiance (1 − q)",    f"{(1-QUANTILE)*100:.0f}%",                              ""),
    ("Date de fin",                    END_DATE,                                                  ""),
    ("Fenêtre VaR (20 ans)",
     f"{ret_var.index[0].date()} → {ret_var.index[-1].date()}",                               f"{n} obs."),
    ("Fenêtre performance",
     f"{ret_perf.index[0].date()} → {ret_perf.index[-1].date()}",                             f"{len(ret_perf)} obs."),
], 4):
    ws4[f"A{i}"].value = lbl; ws4[f"A{i}"].fill = F_LBLUE
    ws4[f"B{i}"].value = v1;  ws4[f"B{i}"].alignment = CTR
    ws4[f"C{i}"].value = v2;  ws4[f"C{i}"].alignment = CTR

# Résultats VaR — référence vers feuille VaR
R0 = 13
ws4.merge_cells(f"A{R0}:C{R0}")
ws4[f"A{R0}"].value = "RÉSULTATS VaR  —  formules liées à la feuille « VaR 95% 1J »"
ws4[f"A{R0}"].fill  = F_RED; ws4[f"A{R0}"].font = font("FFFFFF", bold=True, size=11)
ws4[f"A{R0}"].alignment = LFT

VS = "'VaR 95% 1J'"  # nom feuille entre apostrophes
for i, (lbl, formula, fmt) in enumerate([
    ("n (observations fenêtre VaR)",      f"={VS}!J{RN}",   "0"),
    ("k = FLOOR(n × q, 1)",               f"={VS}!J{RK}",   "0"),
    ("VaR 95% 1J  (%)",                   f"={VS}!J{RVAR}",  "0.0000%"),
    ("VaR 95% 1J  (€)",                   f"={VS}!J{RVAE}",  '#,##0.00 "€"'),
    ("CVaR — Expected Shortfall  (%)",    f"={VS}!J{RCVR}",  "0.0000%"),
    ("CVaR  (€)",                         f"={VS}!J{RCVE}",  '#,##0.00 "€"'),
], R0+1):
    ws4[f"A{i}"].value = lbl; ws4[f"A{i}"].fill = F_LRED
    c = ws4[f"B{i}"]
    c.value = formula; c.number_format = fmt
    c.font = FR; c.alignment = CTR; c.border = BRD

# Résultats performance — référence vers feuille Performance
R1 = R0 + 8
ws4.merge_cells(f"A{R1}:C{R1}")
ws4[f"A{R1}"].value = "MÉTRIQUES DE PERFORMANCE  —  formules liées à la feuille « Performance »"
ws4[f"A{R1}"].fill  = F_GREEN; ws4[f"A{R1}"].font = font("FFFFFF", bold=True, size=11)
ws4[f"A{R1}"].alignment = LFT

PS = "'Performance'"
for i, (lbl, formula, fmt) in enumerate([
    ("Rendement total",             f"={PS}!D{LAST_PERF}",              "0.00%"),
    ("Rendement total (€)",         f"={PS}!D{LAST_PERF}*{AMOUNT}",     '#,##0.00 "€"'),
    ("Valeur finale du portefeuille", f"={PS}!C{LAST_PERF}",            '#,##0.00 "€"'),
    ("Volatilité annualisée",       port_vol/100,                        "0.00%"),
    ("Rendement annuel moyen",      port_mean/100,                       "0.00%"),
], R1+1):
    ws4[f"A{i}"].value = lbl; ws4[f"A{i}"].fill = F_LGREEN
    c = ws4[f"B{i}"]
    c.value = formula; c.number_format = fmt
    c.font = FG; c.alignment = CTR; c.border = BRD

# Résumé des formules clés (documentation)
R2 = R1 + 7
ws4.merge_cells(f"A{R2}:C{R2}")
ws4[f"A{R2}"].value = "FORMULES EXCEL UTILISÉES"
ws4[f"A{R2}"].fill  = F_DBLUE; ws4[f"A{R2}"].font = font("FFFFFF", bold=True, size=11)
ws4[f"A{R2}"].alignment = LFT

for i, (formule, desc) in enumerate([
    ("=SMALL($F$4:$F$N, k)",                        "k-ième plus petite valeur = VaR"),
    ("=AVERAGE($H$4:INDEX($H$4:$H$N, k-1))",        "Moyenne des k−1 pires = CVaR"),
    ("=FLOOR(n × q, 1)",                             "Arrondi inférieur = indice k"),
    ("=SMALL($F$4:$F$N, ROW()-3)",                   "Colonne H triée automatiquement"),
    ("=Ci-1*(1+Bi)",                                  "Valeur portefeuille cumulée"),
    ("=Ci/$C$3-1",                                    "Rendement cumulé depuis t=0"),
], R2+1):
    ws4.cell(i, 1).value = formule; ws4.cell(i, 1).font = font("1F4E79", bold=True)
    ws4.cell(i, 1).fill  = F_LGRAY
    ws4.cell(i, 2).value = desc; ws4.cell(i, 2).fill = F_LGRAY
    ws4.cell(i, 2).alignment = LFT

# Ordre des feuilles : Résumé en premier
wb.move_sheet("Résumé", offset=-3)

# Couleurs des onglets
wb["Résumé"].sheet_properties.tabColor      = "1F4E79"
wb["VaR 95% 1J"].sheet_properties.tabColor  = "C00000"
wb["Performance"].sheet_properties.tabColor = "375623"
wb["Corrélation"].sheet_properties.tabColor = "7030A0"

# ── Sauvegarde ────────────────────────────────────────────────────────────────
output = os.path.join(os.path.expanduser("~"), "Documents", "var_AI_PA_AIR_PA_formules.xlsx")
wb.save(output)
print(f"\nFichier sauvegardé : {output}")
